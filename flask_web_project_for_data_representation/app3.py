
import webbrowser
import os
import json
import requests
from newsapi import NewsApiClient
from openpyxl import Workbook, load_workbook
import collections
import matplotlib.pyplot as plt


from flask import Flask, request, render_template
from flask_mysqldb import MySQL


app = Flask(__name__)
app.config['MYSQL_USER'] = 'sql2383132'
app.config['MYSQL_PASSWORD'] = 'qL1!bV7*'
app.config['MYSQL_HOST'] = 'sql2.freemysqlhosting.net'
app.config['MYSQL_DB'] = 'sql2383132'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'
mysql = MySQL(app)




def view_database_records():


    #Port number: 3306

    cur = mysql.connection.cursor()
    cur.execute('''SELECT * FROM headlinetitles''')
    results = cur.fetchall()
    
    return results


    

def receive_text_from_form(text):

    
    #Setting up extractions from News API

    news_keyword = text

    # https://newsapi.org/docs/client-libraries/python
    newsapi = NewsApiClient(api_key='0fb13acc3bc8480eafedb87afa941f7e')


    # /v2/everything
    data = newsapi.get_everything(q=news_keyword)

    jdict = data.get('articles')
    
    my_list = []
    cur = mysql.connection.cursor()
    for row in jdict:
        transfer = row['title'],row['url']
        headline = row['title']
        headline = headline.replace("'", "") #remove/replace inverted commas to avoid SQL errors when passing data to database
        url = row['url']
        my_list.append(transfer)
        
        sqldata = ("INSERT INTO headlinetitles (Title, Url) VALUES ('%s','%s');",str(headline),str(url))
        cur.execute = (sqldata)
    mysql.connection.commit()        
    return my_list




def activate_all(text):

    #Setting up extractions from News API

    news_keyword = text

    # https://newsapi.org/docs/client-libraries/python
    newsapi = NewsApiClient(api_key='0fb13acc3bc8480eafedb87afa941f7e')


    # /v2/everything
    data = newsapi.get_everything(q=news_keyword)

    jdict = data.get('articles')

    #Updating spreadsheet

    Row_Incrementer = 1

    workbook = Workbook()
    sheet = workbook.active
    sheet["A" + str(Row_Incrementer)] = "Headline Title"
    sheet["B" + str(Row_Incrementer)] = "URL"
    workbook.save(filename="analysis.xlsx")
    workbook.close()


    def excel_row_update(headline_title, url):

        global Row_Incrementer
        Row_Incrementer += 1

        workbook = load_workbook(filename="analysis.xlsx")
        sheet = workbook.active

        sheet["A" + str(Row_Incrementer)] = headline_title
        sheet["B" + str(Row_Incrementer)] = url

        workbook.save(filename="analysis.xlsx")
        workbook.close()


 
    
    #Setting up variables before the loop, these handle the number of iterations in the loop and position to publish content to excel

    row = 0
    excel_start_row = 1


    column = 0
    url=""


    


    my_string_for_word_count = "" #this string will be used for doing word count analysis

    #The while loop below handle printing to console, saving to excel, saving to database



    for row in jdict:
        
        #https://docs.python.org/3/tutorial/errors.html
        try:
          

            headline_title = row['title']
            headline_title = headline_title.replace("'", "") #remove/replace inverted commas to avoid SQL errors when passing data to database

            url = row['url']



            my_string_for_word_count = my_string_for_word_count + headline_title #adding one healine at a time to the string for word count analysis
            excel_row_update(headline_title, url) #update spreadsheet


        except UnicodeEncodeError:
            break
        except:
            break

    
    
    #get current working directory so that files can be published and saved to same folder as the python app
    path = os.getcwd()


    
    #this code below does a word count analysis
    #https://www.w3resource.com/python-exercises/string/python-data-type-string-exercise-12.php
    def word_count(str):
        counts = {}
        words = str.split()

        for word in words:
            if word in counts:
                counts[word] += 1
            else:
                counts[word] = 1

        return counts

    counted_words =  word_count(my_string_for_word_count)


    sorted_dict = collections.OrderedDict(counted_words)



    x = sorted_dict

    #https://stackoverflow.com/questions/613183/how-do-i-sort-a-dictionary-by-value
    sorted_x = sorted(x.items(), key=lambda kv: kv[1])


    #https://stackoverflow.com/questions/646644/how-to-get-last-items-of-a-list-in-python
    my_graph_data = list(sorted_x)[-10:]
    my_graph_data_dict = dict(my_graph_data)



    #https://thispointer.com/different-ways-to-remove-a-key-from-dictionary-in-python/
    common_words = ['the',
    'of',
    'and',
    'a',
    'to',
    'in',
    'is',
    'you',
    'that',
    'it',
    'he',
    'was',
    'for',
    'on',
    'are',
    'as',
    'with',
    'his',
    'they',
    'I',
    'at',
    'be',
    'this',
    'have',
    'from',
    'or',
    'one',
    'had',
    'by',
    'word',
    'but',
    'not',
    'what',
    'all',
    'were',
    'we',
    'when',
    'your',
    'can',
    'said',
    'there',
    'use',
    'an',
    'each',
    'which',
    'she',
    'do',
    'how',
    'their',
    'if',
    'will',
    'up',
    'other',
    'about',
    'out',
    'many',
    'then',
    'them',
    'these',
    'so',
    'some',
    'her',
    'would',
    'make',
    'like',
    'him',
    'into',
    'time',
    'has',
    'look',
    'two',
    'more',
    'write',
    'go',
    'see',
    'number',
    'no',
    'way',
    'could',
    'people',
    'my',
    'than',
    'first',
    'been',
    'call',
    'who',
    'its',
    'now',
    'find',
    'long',
    'down',
    'day',
    'did',
    'get',
    'come',
    'made',
    'may',
    'part',
    ]

    for word in common_words:
        try:
            my_graph_data_dict.pop(word)
        except:
            continue

    #https://www.kite.com/python/answers/how-to-plot-a-bar-chart-using-a-dictionary-in-matplotlib-in-python
    a_dictionary = x
    keys = my_graph_data_dict.keys()
    values = my_graph_data_dict.values()

    #https://www.kite.com/python/answers/how-to-rotate-axis-labels-in-matplotlib-in-python
    plt.xticks(rotation=45)
    plt.yticks(rotation=90)
    #https://showmecode.info/matplotlib/bar/change-bar-color/
    plt.bar(keys, values,  color=['red', 'blue', 'purple', 'green', 'lavender'])
    plt.ylabel('Occurences of word')
    plt.title('Most Frequent Words In Headlines')
    #https://www.kite.com/python/answers/how-save-a-matplotlib-plot-as-a-pdf-file-in-python
    plt.savefig("static/images/plot.png")















@app.route('/')




def my_form():
    return render_template('my-form.html')

@app.route('/', methods=['POST'])
def my_form_post():
    text = request.form['text']







    #get current working directory
    path = os.getcwd()

    
    list = receive_text_from_form(text)
    filelocation = "Find yous files here", path, "/analysis.xlsx"
    activate_all(text)
    return render_template("my-form.html", list=list, filelocation=filelocation)




@app.route('/database')
def database():
    databasestuff = str(view_database_records())
    return """ 
    <style>
    table.blueTable {
  border: 1px solid #1C6EA4;
  background-color: #EEEEEE;
  width: 100%;
  text-align: left;
  border-collapse: collapse;
}
table.blueTable td, table.blueTable th {
  border: 1px solid #AAAAAA;
  padding: 3px 2px;
}
table.blueTable tbody td {
  font-size: 13px;
}
table.blueTable tr:nth-child(even) {
  background: #D0E4F5;
}
table.blueTable thead {
  background: #1C6EA4;
  background: -moz-linear-gradient(top, #5592bb 0%, #327cad 66%, #1C6EA4 100%);
  background: -webkit-linear-gradient(top, #5592bb 0%, #327cad 66%, #1C6EA4 100%);
  background: linear-gradient(to bottom, #5592bb 0%, #327cad 66%, #1C6EA4 100%);
  border-bottom: 2px solid #444444;
}
table.blueTable thead th {
  font-size: 15px;
  font-weight: bold;
  color: #FFFFFF;
  border-left: 2px solid #D0E4F5;
}
table.blueTable thead th:first-child {
  border-left: none;
}

table.blueTable tfoot {
  font-size: 14px;
  font-weight: bold;
  color: #FFFFFF;
  background: #D0E4F5;
  background: -moz-linear-gradient(top, #dcebf7 0%, #d4e6f6 66%, #D0E4F5 100%);
  background: -webkit-linear-gradient(top, #dcebf7 0%, #d4e6f6 66%, #D0E4F5 100%);
  background: linear-gradient(to bottom, #dcebf7 0%, #d4e6f6 66%, #D0E4F5 100%);
  border-top: 2px solid #444444;
}
table.blueTable tfoot td {
  font-size: 14px;
}
table.blueTable tfoot .links {
  text-align: right;
}
table.blueTable tfoot .links a{
  display: inline-block;
  background: #1C6EA4;
  color: #FFFFFF;
  padding: 2px 8px;
  border-radius: 5px;
}
</style>   
    
    <table class="blueTable"><thead><tr><th><centre>Database Contents On MySQL Server</centre></th></tr></thead><tbody><tr><td>""" + databasestuff + """ </td></tr></tbody></tr></table>
    <img src="static/images/plot.png" alt="A Matplotlibplot">
    <button onclick="goBack()">Go Back</button>

<script>
function goBack() {
  window.history.back();
}
</script>
    """ 












if __name__ == "__main__":
    app.run()